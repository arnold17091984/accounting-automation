"""
Tests for Fund Request Module

Tests the fund calculator, expense aggregator, balance tracker, and Excel generation.
"""

import tempfile
from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest

from python.fund_request.fund_calculator import (
    FundCalculator,
    FundRequestData,
    FundRequestItem,
    ProjectExpense,
)
from python.fund_request.expense_aggregator import (
    AggregatedExpense,
    ExpenseAggregator,
)
from python.fund_request.fund_balance_tracker import (
    BankBalance,
    BalanceSummary,
    FundBalanceTracker,
)


# =============================================================================
# FundCalculator Tests
# =============================================================================

class TestFundCalculator:
    """Tests for FundCalculator class."""

    @pytest.fixture
    def calculator(self, tmp_path: Path):
        """Create calculator with test config."""
        config_dir = tmp_path / "config"
        config_dir.mkdir()

        # Create minimal config
        config_file = config_dir / "fund_request_config.yaml"
        config_file.write_text("""
payment_schedule:
  payment_dates:
    - day: 5
      label: "1st Half"
      generate_days_before: 2
    - day: 20
      label: "2nd Half"
      generate_days_before: 2

sections:
  A:
    categories:
      - code: salaries
        name: "Salaries"
      - code: rental
        name: "Office Rental"
  B:
    categories:
      - code: consultation
        name: "Consultation Fee"

validation:
  amount_limits:
    min_item_amount: 0.01
    max_item_amount: 100000000
    max_total_amount: 500000000
  warnings:
    - condition: "overall_total > current_fund_balance * 2"
      message: "Total request exceeds 2x current fund balance"
""")

        return FundCalculator(config_dir)

    def test_create_fund_request_basic(self, calculator: FundCalculator):
        """Test basic fund request creation."""
        fund_request = calculator.create_fund_request(
            entity="tours",
            payment_date=date(2026, 2, 5),
            section_a_items=[
                {"description": "Salaries", "amount": 1000000},
                {"description": "Office Rental", "amount": 500000},
            ],
            section_b_items=[
                {"description": "Consultation Fee", "amount": 50000},
            ],
        )

        assert fund_request.entity == "tours"
        assert fund_request.payment_date == date(2026, 2, 5)
        assert len(fund_request.section_a_items) == 2
        assert len(fund_request.section_b_items) == 1
        assert fund_request.section_a_total == Decimal("1500000")
        assert fund_request.section_b_total == Decimal("50000")
        assert fund_request.overall_total == Decimal("1550000")

    def test_create_fund_request_with_balance(self, calculator: FundCalculator):
        """Test fund request with fund balance info."""
        fund_request = calculator.create_fund_request(
            entity="tours",
            payment_date=date(2026, 2, 5),
            section_a_items=[
                {"description": "Salaries", "amount": 1000000},
            ],
            current_fund_balance=Decimal("2000000"),
            project_expenses=[
                {"project_name": "B-Ticket", "amount": 500000},
                {"project_name": "Daimasu", "amount": 300000},
            ],
        )

        assert fund_request.current_fund_balance == Decimal("2000000")
        assert len(fund_request.project_expenses) == 2
        assert fund_request.project_expenses_total == Decimal("800000")
        assert fund_request.remaining_fund == Decimal("1200000")

    def test_period_label_first_half(self, calculator: FundCalculator):
        """Test period label for 1st half of month."""
        fund_request = calculator.create_fund_request(
            entity="tours",
            payment_date=date(2026, 2, 5),
            section_a_items=[{"description": "Test", "amount": 100}],
        )

        assert "1st Half" in fund_request.period_label
        assert "February" in fund_request.period_label

    def test_period_label_second_half(self, calculator: FundCalculator):
        """Test period label for 2nd half of month."""
        fund_request = calculator.create_fund_request(
            entity="tours",
            payment_date=date(2026, 2, 20),
            section_a_items=[{"description": "Test", "amount": 100}],
        )

        assert "2nd Half" in fund_request.period_label

    def test_validate_fund_request_valid(self, calculator: FundCalculator):
        """Test validation of valid fund request."""
        fund_request = calculator.create_fund_request(
            entity="tours",
            payment_date=date(2026, 2, 5),
            section_a_items=[{"description": "Test", "amount": 100}],
        )

        errors = calculator.validate_fund_request(fund_request)
        assert len(errors) == 0

    def test_validate_fund_request_empty(self, calculator: FundCalculator):
        """Test validation fails for empty fund request."""
        fund_request = FundRequestData(
            entity="tours",
            request_date=date.today(),
            payment_date=date(2026, 2, 5),
        )

        errors = calculator.validate_fund_request(fund_request)
        assert any("at least one section" in e.lower() for e in errors)

    def test_validate_fund_request_missing_entity(self, calculator: FundCalculator):
        """Test validation fails for missing entity."""
        fund_request = FundRequestData(
            entity="",
            request_date=date.today(),
            payment_date=date(2026, 2, 5),
        )
        fund_request.section_a_items.append(
            FundRequestItem(description="Test", amount=Decimal("100"), section="A")
        )

        errors = calculator.validate_fund_request(fund_request)
        assert any("entity" in e.lower() for e in errors)

    def test_get_warnings(self, calculator: FundCalculator):
        """Test warning generation."""
        fund_request = calculator.create_fund_request(
            entity="tours",
            payment_date=date(2026, 2, 5),
            section_a_items=[{"description": "Large expense", "amount": 5000000}],
            current_fund_balance=Decimal("1000000"),
        )

        warnings = calculator.get_warnings(fund_request)
        assert any("2x" in w for w in warnings)

    def test_get_next_payment_date(self, calculator: FundCalculator):
        """Test next payment date calculation."""
        # Before 5th should return 5th
        next_date = calculator.get_next_payment_date(date(2026, 2, 3))
        assert next_date == date(2026, 2, 5)

        # After 5th, before 20th should return 20th
        next_date = calculator.get_next_payment_date(date(2026, 2, 10))
        assert next_date == date(2026, 2, 20)

        # After 20th should return 5th of next month
        next_date = calculator.get_next_payment_date(date(2026, 2, 25))
        assert next_date == date(2026, 3, 5)

    def test_get_entity_title(self, calculator: FundCalculator):
        """Test entity title generation."""
        title = calculator.get_entity_title("tours")
        assert "Tours" in title or "FUND REQUEST" in title

        title = calculator.get_entity_title("solaire")
        assert "Solaire" in title


# =============================================================================
# ExpenseAggregator Tests
# =============================================================================

class TestExpenseAggregator:
    """Tests for ExpenseAggregator class."""

    @pytest.fixture
    def aggregator(self, tmp_path: Path):
        """Create aggregator with test config."""
        config_dir = tmp_path / "config"
        config_dir.mkdir()

        config_file = config_dir / "fund_request_config.yaml"
        config_file.write_text("""
sections:
  A:
    categories:
      - code: salaries
        name: "Salaries"
      - code: rental
        name: "Office Rental"
      - code: credit_card
        name: "Credit Card"
  B:
    categories:
      - code: consultation
        name: "Consultation Fee"
      - code: other
        name: "Other"
""")

        return ExpenseAggregator(config_dir)

    def test_aggregate_recurring_expenses(self, aggregator: ExpenseAggregator):
        """Test aggregation of recurring expenses."""
        recurring_data = [
            {
                "id": 1,
                "entity": "tours",
                "description": "BGC Office Rental",
                "category": "rental",
                "amount": 100000,
                "frequency": "monthly",
                "is_active": True,
            },
            {
                "id": 2,
                "entity": "tours",
                "description": "Salaries",
                "category": "salaries",
                "amount": 500000,
                "frequency": "semi-monthly",
                "is_active": True,
            },
            {
                "id": 3,
                "entity": "solaire",  # Different entity
                "description": "Office",
                "category": "rental",
                "amount": 80000,
                "frequency": "monthly",
                "is_active": True,
            },
        ]

        expenses = aggregator.aggregate_recurring_expenses(
            entity="tours",
            payment_date=date(2026, 2, 5),
            recurring_data=recurring_data,
        )

        # Should only include tours expenses
        assert len(expenses) == 2
        entity_expenses = [e for e in expenses if e.category in ["rental", "salaries"]]
        assert len(entity_expenses) == 2

    def test_aggregate_recurring_excludes_inactive(self, aggregator: ExpenseAggregator):
        """Test that inactive recurring expenses are excluded."""
        recurring_data = [
            {
                "id": 1,
                "entity": "tours",
                "description": "Active expense",
                "category": "rental",
                "amount": 100000,
                "frequency": "monthly",
                "is_active": True,
            },
            {
                "id": 2,
                "entity": "tours",
                "description": "Inactive expense",
                "category": "rental",
                "amount": 50000,
                "frequency": "monthly",
                "is_active": False,
            },
        ]

        expenses = aggregator.aggregate_recurring_expenses(
            entity="tours",
            payment_date=date(2026, 2, 5),
            recurring_data=recurring_data,
        )

        assert len(expenses) == 1
        assert expenses[0].description == "Active expense"

    def test_aggregate_transactions(self, aggregator: ExpenseAggregator):
        """Test aggregation of transactions."""
        transactions = [
            {
                "id": "1",
                "entity": "tours",
                "txn_date": "2026-02-01",
                "category": "expense",
                "amount": 1000,
            },
            {
                "id": "2",
                "entity": "tours",
                "txn_date": "2026-02-02",
                "category": "expense",
                "amount": 2000,
            },
            {
                "id": "3",
                "entity": "tours",
                "txn_date": "2026-02-03",
                "category": "travel",
                "amount": 500,
            },
        ]

        expenses = aggregator.aggregate_transactions(
            entity="tours",
            start_date=date(2026, 2, 1),
            end_date=date(2026, 2, 28),
            transactions=transactions,
            group_by="category",
        )

        assert len(expenses) == 2

        expense_cat = next((e for e in expenses if e.category == "expense"), None)
        assert expense_cat is not None
        assert expense_cat.amount == Decimal("3000")
        assert expense_cat.source_count == 2

    def test_categorize_to_section(self, aggregator: ExpenseAggregator):
        """Test category to section mapping."""
        assert aggregator.categorize_to_section("salaries") == "A"
        assert aggregator.categorize_to_section("rental") == "A"
        assert aggregator.categorize_to_section("consultation") == "B"
        assert aggregator.categorize_to_section("unknown") == "B"  # Default

    def test_sort_expenses_by_priority(self, aggregator: ExpenseAggregator):
        """Test expense sorting by priority."""
        expenses = [
            AggregatedExpense(category="other", description="Other", amount=Decimal("100")),
            AggregatedExpense(category="rental", description="Rental", amount=Decimal("200")),
            AggregatedExpense(category="salaries", description="Salaries", amount=Decimal("300")),
        ]

        sorted_expenses = aggregator.sort_expenses_by_priority(expenses)

        assert sorted_expenses[0].category == "rental"
        assert sorted_expenses[1].category == "salaries"
        assert sorted_expenses[2].category == "other"


# =============================================================================
# FundBalanceTracker Tests
# =============================================================================

class TestFundBalanceTracker:
    """Tests for FundBalanceTracker class."""

    @pytest.fixture
    def tracker(self, tmp_path: Path):
        """Create tracker with test config."""
        config_dir = tmp_path / "config"
        config_dir.mkdir()

        config_file = config_dir / "entity_config.yaml"
        config_file.write_text("""
entity_overrides:
  tours:
    projects:
      - "B-Ticket"
      - "Daimasu"
reference_info:
  default_projects:
    - "Project A"
    - "Project B"
""")

        return FundBalanceTracker(config_dir)

    def test_get_current_balance(self, tracker: FundBalanceTracker):
        """Test getting current balance for entity."""
        balance_records = [
            {
                "entity": "tours",
                "account_name": "Main Fund",
                "balance": 1000000,
                "balance_date": "2026-02-01",
            },
            {
                "entity": "tours",
                "account_name": "Main Fund",
                "balance": 1500000,
                "balance_date": "2026-02-05",  # Later date
            },
            {
                "entity": "tours",
                "account_name": "Petty Cash",
                "balance": 50000,
                "balance_date": "2026-02-03",
            },
        ]

        summary = tracker.get_current_balance("tours", balance_records)

        assert summary.entity == "tours"
        assert summary.total_balance == Decimal("1550000")  # Latest Main Fund + Petty Cash
        assert len(summary.accounts) == 2

    def test_calculate_remaining_fund(self, tracker: FundBalanceTracker):
        """Test remaining fund calculation."""
        remaining = tracker.calculate_remaining_fund(
            current_balance=Decimal("2000000"),
            project_expenses=[
                {"amount": 500000},
                {"amount": 300000},
            ],
        )

        assert remaining == Decimal("1200000")

    def test_record_balance(self, tracker: FundBalanceTracker):
        """Test balance record creation."""
        balance = tracker.record_balance(
            entity="tours",
            account_name="Main Fund",
            balance=Decimal("1500000"),
            balance_date=date(2026, 2, 5),
            bank="unionbank",
            source="manual",
        )

        assert balance.entity == "tours"
        assert balance.account_name == "Main Fund"
        assert balance.balance == Decimal("1500000")
        assert balance.bank == "unionbank"

    def test_validate_balance_entry_negative(self, tracker: FundBalanceTracker):
        """Test validation warns on negative balance."""
        balance = BankBalance(
            entity="tours",
            account_name="Main Fund",
            balance=Decimal("-1000"),
        )

        warnings = tracker.validate_balance_entry(balance)
        assert any("negative" in w.lower() for w in warnings)

    def test_validate_balance_entry_large_change(self, tracker: FundBalanceTracker):
        """Test validation warns on large balance change."""
        previous = BankBalance(
            entity="tours",
            account_name="Main Fund",
            balance=Decimal("1000000"),
        )

        current = BankBalance(
            entity="tours",
            account_name="Main Fund",
            balance=Decimal("100000"),  # 90% decrease
        )

        warnings = tracker.validate_balance_entry(current, previous)
        assert any("large" in w.lower() for w in warnings)


# =============================================================================
# FundRequestData Tests
# =============================================================================

class TestFundRequestData:
    """Tests for FundRequestData dataclass."""

    def test_calculate_totals(self):
        """Test total calculation."""
        data = FundRequestData(
            entity="tours",
            request_date=date.today(),
            payment_date=date(2026, 2, 5),
        )

        data.section_a_items = [
            FundRequestItem(description="Item 1", amount=Decimal("1000"), section="A"),
            FundRequestItem(description="Item 2", amount=Decimal("2000"), section="A"),
        ]
        data.section_b_items = [
            FundRequestItem(description="Item 3", amount=Decimal("500"), section="B"),
        ]

        data.calculate_totals()

        assert data.section_a_total == Decimal("3000")
        assert data.section_b_total == Decimal("500")
        assert data.overall_total == Decimal("3500")

    def test_add_section_items(self):
        """Test adding items to sections."""
        data = FundRequestData(
            entity="tours",
            request_date=date.today(),
            payment_date=date(2026, 2, 5),
        )

        item_a = FundRequestItem(description="Test A", amount=Decimal("1000"), section="")
        data.add_section_a_item(item_a)

        item_b = FundRequestItem(description="Test B", amount=Decimal("500"), section="")
        data.add_section_b_item(item_b)

        assert len(data.section_a_items) == 1
        assert data.section_a_items[0].section == "A"
        assert data.section_a_items[0].line_number == 1

        assert len(data.section_b_items) == 1
        assert data.section_b_items[0].section == "B"

        assert data.overall_total == Decimal("1500")

    def test_to_dict(self):
        """Test conversion to dictionary."""
        data = FundRequestData(
            entity="tours",
            request_date=date(2026, 2, 3),
            payment_date=date(2026, 2, 5),
            period_label="February 2026 - 1st Half",
        )
        data.section_a_items.append(
            FundRequestItem(description="Test", amount=Decimal("1000"), section="A")
        )
        data.calculate_totals()

        result = data.to_dict()

        assert result["entity"] == "tours"
        assert result["payment_date"] == "2026-02-05"
        assert result["overall_total"] == 1000.0
        assert len(result["section_a_items"]) == 1


# =============================================================================
# Excel Generator Tests (requires openpyxl)
# =============================================================================

class TestExcelGenerator:
    """Tests for FundRequestExcelGenerator."""

    @pytest.fixture
    def generator(self, tmp_path: Path):
        """Create generator with test config."""
        try:
            from python.fund_request.excel_generator import FundRequestExcelGenerator
        except ImportError:
            pytest.skip("openpyxl not installed")

        config_dir = tmp_path / "config"
        config_dir.mkdir()

        config_file = config_dir / "fund_request_config.yaml"
        config_file.write_text("""
excel_template:
  filename_pattern: "{entity}_FundRequest_{date}.xlsx"
  sheet_name: "Fund Request"
  styles:
    header:
      font: "Arial"
      font_size: 14
      bold: true
      fill_color: "FFFF00"
    currency_format: "#,##0.00"
  columns:
    A: { width: 5 }
    B: { width: 40 }
    C: { width: 15 }
entity_overrides:
  tours:
    section_title: "Betrnk Tours FUND REQUEST"
    reference_recipient: "Kent-san"
""")

        return FundRequestExcelGenerator(config_dir)

    def test_generate_excel_basic(self, generator, tmp_path: Path):
        """Test basic Excel generation."""
        from python.fund_request.fund_calculator import FundCalculator

        calculator = FundCalculator(generator.config_dir)
        fund_request = calculator.create_fund_request(
            entity="tours",
            payment_date=date(2026, 2, 5),
            section_a_items=[
                {"description": "Salaries", "amount": 1000000},
                {"description": "Rental", "amount": 500000},
            ],
            section_b_items=[
                {"description": "Consultation", "amount": 50000},
            ],
        )

        output_path = tmp_path / "test_output.xlsx"
        result_path = generator.generate(fund_request, output_path)

        assert result_path.exists()
        assert result_path.suffix == ".xlsx"

    def test_generate_excel_with_reference_info(self, generator, tmp_path: Path):
        """Test Excel generation with reference info."""
        from python.fund_request.fund_calculator import FundCalculator

        calculator = FundCalculator(generator.config_dir)
        fund_request = calculator.create_fund_request(
            entity="tours",
            payment_date=date(2026, 2, 5),
            section_a_items=[
                {"description": "Salaries", "amount": 1000000},
            ],
            current_fund_balance=Decimal("2000000"),
            project_expenses=[
                {"project_name": "B-Ticket", "amount": 500000},
                {"project_name": "Daimasu", "amount": 300000},
            ],
        )

        output_path = tmp_path / "test_with_ref.xlsx"
        result_path = generator.generate(fund_request, output_path)

        assert result_path.exists()

        # Verify content by reading the file
        from openpyxl import load_workbook
        wb = load_workbook(result_path)
        ws = wb.active

        # Check that reference info is included
        found_balance = False
        found_project = False
        for row in ws.iter_rows(values_only=True):
            row_text = " ".join(str(c) for c in row if c)
            if "Current Fund Balance" in row_text:
                found_balance = True
            if "B-Ticket" in row_text:
                found_project = True

        assert found_balance, "Current fund balance not found in Excel"
        assert found_project, "Project expenses not found in Excel"

    def test_generate_excel_auto_path(self, generator):
        """Test Excel generation with auto-generated path."""
        from python.fund_request.fund_calculator import FundCalculator

        calculator = FundCalculator(generator.config_dir)
        fund_request = calculator.create_fund_request(
            entity="tours",
            payment_date=date(2026, 2, 5),
            section_a_items=[{"description": "Test", "amount": 100}],
        )

        result_path = generator.generate(fund_request)

        assert result_path.exists()
        assert "Tours" in result_path.name
        assert "FundRequest" in result_path.name

        # Clean up
        result_path.unlink()
