"""
Fund Request Excel Generator Module

Generates formatted Excel documents for fund requests.
"""

import logging
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Any

import yaml

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed. Excel generation will not work.")

from .fund_calculator import FundRequestData


class FundRequestExcelGenerator:
    """Generates Excel files for fund requests."""

    def __init__(self, config_dir: Path | str | None = None):
        """Initialize the generator.

        Args:
            config_dir: Path to configuration directory
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel generation")

        self.config_dir = Path(config_dir) if config_dir else Path(__file__).parent.parent.parent / "config"
        self._load_config()
        self._setup_styles()

    def _load_config(self) -> None:
        """Load configuration from YAML."""
        config_file = self.config_dir / "fund_request_config.yaml"
        if config_file.exists():
            with open(config_file) as f:
                self.config = yaml.safe_load(f)
        else:
            self.config = {}

        self.template_config = self.config.get("excel_template", {})

    def _setup_styles(self) -> None:
        """Setup Excel styles from config."""
        styles = self.template_config.get("styles", {})

        # Header style (title)
        header_config = styles.get("header", {})
        self.header_font = Font(
            name=header_config.get("font", "Arial"),
            size=header_config.get("font_size", 14),
            bold=header_config.get("bold", True),
        )
        self.header_fill = PatternFill(
            start_color=header_config.get("fill_color", "FFFF00"),
            end_color=header_config.get("fill_color", "FFFF00"),
            fill_type="solid",
        )

        # Section header style
        section_config = styles.get("section_header", {})
        self.section_font = Font(
            name=section_config.get("font", "Arial"),
            size=section_config.get("font_size", 11),
            bold=section_config.get("bold", True),
        )
        self.section_fill = PatternFill(
            start_color=section_config.get("fill_color", "D9E1F2"),
            end_color=section_config.get("fill_color", "D9E1F2"),
            fill_type="solid",
        )

        # Total row style
        total_config = styles.get("total_row", {})
        self.total_font = Font(
            name=total_config.get("font", "Arial"),
            size=total_config.get("font_size", 11),
            bold=total_config.get("bold", True),
        )
        self.total_fill = PatternFill(
            start_color=total_config.get("fill_color", "FFFF00"),
            end_color=total_config.get("fill_color", "FFFF00"),
            fill_type="solid",
        )

        # Normal style
        self.normal_font = Font(name="Arial", size=10)

        # Border style
        thin_border = Side(style="thin", color="000000")
        self.border = Border(
            left=thin_border,
            right=thin_border,
            top=thin_border,
            bottom=thin_border,
        )

        # Alignments
        self.center_align = Alignment(horizontal="center", vertical="center")
        self.left_align = Alignment(horizontal="left", vertical="center")
        self.right_align = Alignment(horizontal="right", vertical="center")

        # Number format
        self.currency_format = styles.get("currency_format", "#,##0.00")

    def generate(
        self,
        fund_request: FundRequestData,
        output_path: Path | str | None = None,
    ) -> Path:
        """Generate Excel file for fund request.

        Args:
            fund_request: FundRequestData to generate
            output_path: Output file path (generated if None)

        Returns:
            Path to generated file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = self.template_config.get("sheet_name", "Fund Request")

        current_row = 1

        # Set column widths
        self._set_column_widths(ws)

        # Title
        current_row = self._write_title(ws, current_row, fund_request)

        # Payment date
        current_row = self._write_date_info(ws, current_row, fund_request)

        # Empty row
        current_row += 1

        # Section A - Regular Expenses
        current_row = self._write_section_a(ws, current_row, fund_request)

        # Empty row
        current_row += 1

        # Section B - Others
        current_row = self._write_section_b(ws, current_row, fund_request)

        # Empty row
        current_row += 1

        # Overall Total
        current_row = self._write_overall_total(ws, current_row, fund_request)

        # Empty row
        current_row += 1

        # Reference info (if available)
        if fund_request.current_fund_balance is not None:
            current_row = self._write_reference_info(ws, current_row, fund_request)

        # Generate output path
        if output_path is None:
            output_path = self._generate_output_path(fund_request)
        else:
            output_path = Path(output_path)

        # Ensure directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Save workbook
        wb.save(output_path)
        logger.info(f"Generated fund request Excel: {output_path}")

        return output_path

    def _set_column_widths(self, ws) -> None:
        """Set column widths from config."""
        columns = self.template_config.get("columns", {})

        default_widths = {
            "A": 5,
            "B": 40,
            "C": 15,
            "D": 15,
            "E": 20,
        }

        for col_letter, config in columns.items():
            width = config.get("width", default_widths.get(col_letter, 10))
            ws.column_dimensions[col_letter].width = width

        # Set remaining columns to default
        for col_letter, width in default_widths.items():
            if col_letter not in columns:
                ws.column_dimensions[col_letter].width = width

    def _write_title(self, ws, row: int, fund_request: FundRequestData) -> int:
        """Write title row."""
        from .fund_calculator import FundCalculator
        calculator = FundCalculator(self.config_dir)
        title = calculator.get_entity_title(fund_request.entity)

        # Merge cells for title
        ws.merge_cells(f"A{row}:E{row}")
        cell = ws[f"A{row}"]
        cell.value = title
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.alignment = self.center_align

        return row + 1

    def _write_date_info(self, ws, row: int, fund_request: FundRequestData) -> int:
        """Write date information."""
        # Period label
        if fund_request.period_label:
            ws.merge_cells(f"A{row}:E{row}")
            cell = ws[f"A{row}"]
            cell.value = fund_request.period_label
            cell.font = self.normal_font
            cell.alignment = self.center_align
            row += 1

        # Payment date
        ws.merge_cells(f"A{row}:E{row}")
        cell = ws[f"A{row}"]
        cell.value = f"Payment Date: {fund_request.payment_date.strftime('%B %d, %Y')}"
        cell.font = self.normal_font
        cell.alignment = self.center_align

        return row + 1

    def _write_section_a(self, ws, row: int, fund_request: FundRequestData) -> int:
        """Write Section A (Regular Expenses)."""
        # Section header
        ws.merge_cells(f"A{row}:E{row}")
        cell = ws[f"A{row}"]
        cell.value = "A. Regular Expenses"
        cell.font = self.section_font
        cell.fill = self.section_fill
        row += 1

        # Column headers
        headers = ["#", "Description", "Amount (PHP)", "Amount (USD)", "Notes"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.section_font
            cell.border = self.border
            cell.alignment = self.center_align
        row += 1

        # Items
        for item in fund_request.section_a_items:
            ws.cell(row=row, column=1, value=item.line_number).border = self.border
            ws.cell(row=row, column=2, value=item.description).border = self.border

            amount_cell = ws.cell(row=row, column=3, value=float(item.amount))
            amount_cell.number_format = self.currency_format
            amount_cell.border = self.border
            amount_cell.alignment = self.right_align

            # USD column (if currency is USD)
            usd_cell = ws.cell(row=row, column=4, value="")
            if item.currency == "USD":
                usd_cell.value = float(item.amount)
                usd_cell.number_format = self.currency_format
            usd_cell.border = self.border

            ws.cell(row=row, column=5, value=item.notes or "").border = self.border
            row += 1

        # Section total
        ws.cell(row=row, column=1, value="").border = self.border
        total_label = ws.cell(row=row, column=2, value="Sub-Total Section A")
        total_label.font = self.total_font
        total_label.border = self.border

        total_cell = ws.cell(row=row, column=3, value=float(fund_request.section_a_total))
        total_cell.font = self.total_font
        total_cell.fill = self.total_fill
        total_cell.number_format = self.currency_format
        total_cell.border = self.border
        total_cell.alignment = self.right_align

        ws.cell(row=row, column=4, value="").border = self.border
        ws.cell(row=row, column=5, value="").border = self.border

        return row + 1

    def _write_section_b(self, ws, row: int, fund_request: FundRequestData) -> int:
        """Write Section B (Others)."""
        # Section header
        ws.merge_cells(f"A{row}:E{row}")
        cell = ws[f"A{row}"]
        cell.value = "B. Others"
        cell.font = self.section_font
        cell.fill = self.section_fill
        row += 1

        # Column headers
        headers = ["#", "Description", "Amount (PHP)", "Amount (USD)", "Notes"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.section_font
            cell.border = self.border
            cell.alignment = self.center_align
        row += 1

        # Items
        for item in fund_request.section_b_items:
            ws.cell(row=row, column=1, value=item.line_number).border = self.border
            ws.cell(row=row, column=2, value=item.description).border = self.border

            amount_cell = ws.cell(row=row, column=3, value=float(item.amount))
            amount_cell.number_format = self.currency_format
            amount_cell.border = self.border
            amount_cell.alignment = self.right_align

            usd_cell = ws.cell(row=row, column=4, value="")
            if item.currency == "USD":
                usd_cell.value = float(item.amount)
                usd_cell.number_format = self.currency_format
            usd_cell.border = self.border

            ws.cell(row=row, column=5, value=item.notes or "").border = self.border
            row += 1

        # Section total
        ws.cell(row=row, column=1, value="").border = self.border
        total_label = ws.cell(row=row, column=2, value="Sub-Total Section B")
        total_label.font = self.total_font
        total_label.border = self.border

        total_cell = ws.cell(row=row, column=3, value=float(fund_request.section_b_total))
        total_cell.font = self.total_font
        total_cell.fill = self.total_fill
        total_cell.number_format = self.currency_format
        total_cell.border = self.border
        total_cell.alignment = self.right_align

        ws.cell(row=row, column=4, value="").border = self.border
        ws.cell(row=row, column=5, value="").border = self.border

        return row + 1

    def _write_overall_total(self, ws, row: int, fund_request: FundRequestData) -> int:
        """Write overall total row."""
        ws.merge_cells(f"A{row}:B{row}")
        label_cell = ws[f"A{row}"]
        label_cell.value = "OVERALL TOTAL"
        label_cell.font = Font(name="Arial", size=12, bold=True)
        label_cell.fill = self.total_fill
        label_cell.alignment = self.center_align
        label_cell.border = self.border
        ws[f"B{row}"].border = self.border

        total_cell = ws.cell(row=row, column=3, value=float(fund_request.overall_total))
        total_cell.font = Font(name="Arial", size=12, bold=True)
        total_cell.fill = self.total_fill
        total_cell.number_format = self.currency_format
        total_cell.border = self.border
        total_cell.alignment = self.right_align

        ws.cell(row=row, column=4, value="").border = self.border
        ws.cell(row=row, column=5, value="").border = self.border

        return row + 1

    def _write_reference_info(self, ws, row: int, fund_request: FundRequestData) -> int:
        """Write reference info section."""
        # Get reference recipient from config
        overrides = self.config.get("entity_overrides", {})
        entity_override = overrides.get(fund_request.entity, {})
        recipient = entity_override.get("reference_recipient", "")

        # Section header
        ws.merge_cells(f"A{row}:E{row}")
        cell = ws[f"A{row}"]
        title = f"Reference Information"
        if recipient:
            title += f" (for {recipient})"
        cell.value = title
        cell.font = self.section_font
        cell.fill = self.section_fill
        row += 1

        # Current fund balance
        ws.cell(row=row, column=1, value="")
        ws.cell(row=row, column=2, value="Current Fund Balance:")
        balance_cell = ws.cell(row=row, column=3, value=float(fund_request.current_fund_balance))
        balance_cell.number_format = self.currency_format
        balance_cell.alignment = self.right_align
        row += 1

        # Project expenses header
        if fund_request.project_expenses:
            row += 1
            ws.cell(row=row, column=2, value="Project Expenses:").font = self.section_font
            row += 1

            for pe in fund_request.project_expenses:
                ws.cell(row=row, column=2, value=f"  {pe.project_name}")
                pe_cell = ws.cell(row=row, column=3, value=float(pe.amount))
                pe_cell.number_format = self.currency_format
                pe_cell.alignment = self.right_align
                row += 1

            # Project total
            ws.cell(row=row, column=2, value="Project Expenses Total:").font = self.total_font
            total_cell = ws.cell(row=row, column=3, value=float(fund_request.project_expenses_total))
            total_cell.font = self.total_font
            total_cell.number_format = self.currency_format
            total_cell.alignment = self.right_align
            row += 1

        # Remaining fund
        if fund_request.remaining_fund is not None:
            row += 1
            ws.cell(row=row, column=2, value="Remaining Fund:").font = self.total_font
            remaining_cell = ws.cell(row=row, column=3, value=float(fund_request.remaining_fund))
            remaining_cell.font = self.total_font
            remaining_cell.number_format = self.currency_format
            remaining_cell.alignment = self.right_align

            # Highlight if negative
            if fund_request.remaining_fund < 0:
                remaining_cell.font = Font(name="Arial", size=11, bold=True, color="FF0000")

        return row + 1

    def _generate_output_path(self, fund_request: FundRequestData) -> Path:
        """Generate output file path based on config pattern."""
        pattern = self.template_config.get("filename_pattern", "{entity}_FundRequest_{date}.xlsx")

        filename = pattern.format(
            entity=fund_request.entity.title(),
            date=fund_request.request_date.strftime("%Y%m%d"),
        )

        # Default output directory
        output_dir = Path(__file__).parent / "output"
        output_dir.mkdir(parents=True, exist_ok=True)

        return output_dir / filename


def generate_fund_request_excel(
    fund_request: FundRequestData,
    output_path: Path | str | None = None,
    config_dir: Path | str | None = None,
) -> Path:
    """Convenience function to generate fund request Excel.

    Args:
        fund_request: FundRequestData to generate
        output_path: Output file path
        config_dir: Configuration directory

    Returns:
        Path to generated file
    """
    generator = FundRequestExcelGenerator(config_dir)
    return generator.generate(fund_request, output_path)
