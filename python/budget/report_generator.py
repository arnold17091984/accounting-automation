"""
Budget Report Generator Module

Generates formatted budget reports for various channels (Telegram, Excel, etc).
"""

import json
import logging
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .variance_calculator import VarianceItem, VarianceReport
from .threshold_checker import ThresholdAlert, ThresholdCheckResult
from .historical_analyzer import BudgetAnalysisResult, BudgetSuggestion

logger = logging.getLogger(__name__)


class BudgetReportGenerator:
    """Generates budget reports in various formats."""

    # Formatting constants
    CURRENCY_FORMAT = "₱{:,.2f}"
    PERCENT_FORMAT = "{:.1f}%"

    # Status emojis
    STATUS_EMOJI = {
        "ok": "✅",
        "warning": "⚠️",
        "critical": "🟠",
        "exceeded": "🔴"
    }

    SEVERITY_EMOJI = {
        "low": "⚠️",
        "medium": "🟠",
        "high": "🔴"
    }

    def __init__(self, config_dir: Path | str | None = None):
        """Initialize the report generator.

        Args:
            config_dir: Path to configuration directory
        """
        self.config_dir = Path(config_dir) if config_dir else None

    def format_variance_for_telegram(
        self,
        report: VarianceReport,
        include_all: bool = False,
        max_items: int = 10
    ) -> str:
        """Format variance report for Telegram.

        Args:
            report: VarianceReport to format
            include_all: Include all items (not just warnings)
            max_items: Maximum items to include

        Returns:
            Formatted message string
        """
        lines = []

        # Header
        entity_display = report.entity.upper() if report.entity else "ALL ENTITIES"
        lines.append(f"📊 *Budget Status Report*")
        lines.append(f"Entity: {entity_display}")
        lines.append(f"Period: {report.period}")
        lines.append("")

        # Summary
        summary = report.summary
        lines.append("*Summary:*")
        lines.append(f"• Total Budget: {self.CURRENCY_FORMAT.format(summary.get('total_budget', 0))}")
        lines.append(f"• Total Spent: {self.CURRENCY_FORMAT.format(summary.get('total_actual', 0))}")
        lines.append(f"• Utilization: {self.PERCENT_FORMAT.format(summary.get('overall_utilization', 0))}")
        lines.append("")

        # Status counts
        lines.append("*Status:*")
        lines.append(f"✅ OK: {summary.get('accounts_ok', 0)}")
        lines.append(f"⚠️ Warning/Critical: {summary.get('accounts_warning', 0)}")
        lines.append(f"🔴 Over Budget: {summary.get('accounts_over_budget', 0)}")
        lines.append("")

        # Items needing attention
        warning_items = [i for i in report.items if i.status in ["warning", "critical", "exceeded"]]

        if warning_items:
            lines.append("*Items Needing Attention:*")

            for item in sorted(warning_items, key=lambda x: x.utilization_percent, reverse=True)[:max_items]:
                emoji = self.STATUS_EMOJI.get(item.status, "")
                lines.append(
                    f"{emoji} {item.account_name}: {self.PERCENT_FORMAT.format(item.utilization_percent)} "
                    f"({self.CURRENCY_FORMAT.format(float(item.actual_amount))} / "
                    f"{self.CURRENCY_FORMAT.format(float(item.budget_amount))})"
                )

            if len(warning_items) > max_items:
                lines.append(f"_...and {len(warning_items) - max_items} more_")

        elif include_all:
            lines.append("*All Accounts:*")
            for item in sorted(report.items, key=lambda x: x.utilization_percent, reverse=True)[:max_items]:
                emoji = self.STATUS_EMOJI.get(item.status, "✅")
                lines.append(f"{emoji} {item.account_name}: {self.PERCENT_FORMAT.format(item.utilization_percent)}")
        else:
            lines.append("✅ All accounts within budget!")

        lines.append("")
        lines.append(f"_Generated: {report.generated_at.strftime('%Y-%m-%d %H:%M')}_")

        return "\n".join(lines)

    def format_alerts_for_telegram(
        self,
        alerts: list[ThresholdAlert],
        title: str = "Budget Alerts"
    ) -> str:
        """Format threshold alerts for Telegram.

        Args:
            alerts: List of alerts to format
            title: Title for the message

        Returns:
            Formatted message string
        """
        if not alerts:
            return "✅ No budget alerts at this time."

        lines = [f"🚨 *{title}*", ""]

        # Group by severity
        high = [a for a in alerts if a.severity == "high"]
        medium = [a for a in alerts if a.severity == "medium"]
        low = [a for a in alerts if a.severity == "low"]

        if high:
            lines.append("*🔴 Critical (Over Budget):*")
            for alert in high:
                lines.append(
                    f"• {alert.entity} - {alert.account_name}: "
                    f"{self.PERCENT_FORMAT.format(alert.actual_percent)} "
                    f"(₱{float(alert.actual_amount):,.0f} / ₱{float(alert.budget_amount):,.0f})"
                )
            lines.append("")

        if medium:
            lines.append("*🟠 Warning (>90%):*")
            for alert in medium:
                lines.append(
                    f"• {alert.entity} - {alert.account_name}: "
                    f"{self.PERCENT_FORMAT.format(alert.actual_percent)}"
                )
            lines.append("")

        if low:
            lines.append("*⚠️ Approaching (>70%):*")
            for alert in low[:5]:  # Limit low severity
                lines.append(
                    f"• {alert.entity} - {alert.account_name}: "
                    f"{self.PERCENT_FORMAT.format(alert.actual_percent)}"
                )
            if len(low) > 5:
                lines.append(f"_...and {len(low) - 5} more_")

        return "\n".join(lines)

    def format_suggestions_for_telegram(
        self,
        result: BudgetAnalysisResult,
        max_items: int = 8
    ) -> str:
        """Format budget suggestions for Telegram.

        Args:
            result: BudgetAnalysisResult to format
            max_items: Maximum items to show

        Returns:
            Formatted message string
        """
        lines = [
            f"📈 *Budget Recommendations*",
            f"Entity: {result.entity.upper()}",
            f"Target: {result.target_period}",
            ""
        ]

        # Summary
        change = result.total_change_percent
        change_emoji = "📈" if change > 0 else "📉" if change < 0 else "➡️"
        lines.append(
            f"*Total: {self.CURRENCY_FORMAT.format(float(result.total_current))} → "
            f"{self.CURRENCY_FORMAT.format(float(result.total_recommended))}* "
            f"({change_emoji} {change:+.1f}%)"
        )
        lines.append("")

        # Top changes
        significant = [s for s in result.suggestions if abs(s.change_percent) > 10]
        significant.sort(key=lambda x: abs(x.change_percent), reverse=True)

        if significant:
            lines.append("*Significant Changes:*")
            for s in significant[:max_items]:
                emoji = "⬆️" if s.change_percent > 0 else "⬇️"
                lines.append(
                    f"{emoji} {s.account_name}: {s.change_percent:+.0f}% "
                    f"(→ {self.CURRENCY_FORMAT.format(float(s.recommended_budget))})"
                )
            lines.append("")

        # Key insights
        if result.key_insights:
            lines.append("*Key Insights:*")
            for insight in result.key_insights[:3]:
                lines.append(f"• {insight}")
            lines.append("")

        # Risks
        high_risk = [s for s in result.suggestions if s.risk_level == "high"]
        if high_risk:
            lines.append("⚠️ *High Variability Accounts:*")
            for s in high_risk[:3]:
                lines.append(f"• {s.account_name}")

        lines.append("")
        lines.append(f"_Analysis date: {result.analysis_date.strftime('%Y-%m-%d')}_")

        return "\n".join(lines)

    def generate_variance_excel(
        self,
        report: VarianceReport,
        output_path: Path | str
    ) -> Path:
        """Generate Excel variance report.

        Args:
            report: VarianceReport to export
            output_path: Output file path

        Returns:
            Path to generated file
        """
        output_path = Path(output_path)
        wb = Workbook()
        ws = wb.active
        ws.title = "Budget Variance"

        # Styles
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        warning_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        critical_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        exceeded_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = f"Budget Variance Report - {report.entity or 'All Entities'} - {report.period}"
        ws['A1'].font = Font(bold=True, size=14)

        # Headers
        headers = ['Entity', 'Account', 'Name', 'Budget', 'Actual', 'Variance', 'Utilization', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Data
        for row_num, item in enumerate(report.items, 4):
            ws.cell(row=row_num, column=1, value=item.entity)
            ws.cell(row=row_num, column=2, value=item.account_code)
            ws.cell(row=row_num, column=3, value=item.account_name)
            ws.cell(row=row_num, column=4, value=float(item.budget_amount))
            ws.cell(row=row_num, column=5, value=float(item.actual_amount))
            ws.cell(row=row_num, column=6, value=float(item.variance_amount))
            ws.cell(row=row_num, column=7, value=item.utilization_percent / 100)
            ws.cell(row=row_num, column=8, value=item.status.upper())

            # Format currency columns
            for col in [4, 5, 6]:
                ws.cell(row=row_num, column=col).number_format = '₱#,##0.00'
            ws.cell(row=row_num, column=7).number_format = '0.0%'

            # Conditional formatting
            if item.status == "exceeded":
                for col in range(1, 9):
                    ws.cell(row=row_num, column=col).fill = exceeded_fill
            elif item.status == "critical":
                for col in range(1, 9):
                    ws.cell(row=row_num, column=col).fill = critical_fill
            elif item.status == "warning":
                for col in range(1, 9):
                    ws.cell(row=row_num, column=col).fill = warning_fill

        # Auto-fit columns
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)

        # Summary section
        summary_row = len(report.items) + 6
        ws.cell(row=summary_row, column=1, value="Summary")
        ws.cell(row=summary_row, column=1).font = Font(bold=True)

        ws.cell(row=summary_row + 1, column=1, value="Total Budget")
        ws.cell(row=summary_row + 1, column=2, value=report.summary.get('total_budget', 0))
        ws.cell(row=summary_row + 1, column=2).number_format = '₱#,##0.00'

        ws.cell(row=summary_row + 2, column=1, value="Total Actual")
        ws.cell(row=summary_row + 2, column=2, value=report.summary.get('total_actual', 0))
        ws.cell(row=summary_row + 2, column=2).number_format = '₱#,##0.00'

        ws.cell(row=summary_row + 3, column=1, value="Overall Utilization")
        ws.cell(row=summary_row + 3, column=2, value=report.summary.get('overall_utilization', 0) / 100)
        ws.cell(row=summary_row + 3, column=2).number_format = '0.0%'

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        return output_path

    def format_daily_digest(
        self,
        entity: str,
        date: str,
        transactions_today: int,
        total_spent: float,
        mtd_budget: float,
        mtd_actual: float,
        top_categories: list[tuple[str, float]]
    ) -> str:
        """Format a daily budget digest for Telegram.

        Args:
            entity: Entity code
            date: Date string
            transactions_today: Number of transactions
            total_spent: Total spent today
            mtd_budget: Month-to-date budget
            mtd_actual: Month-to-date actual
            top_categories: List of (category, amount) tuples

        Returns:
            Formatted message string
        """
        utilization = (mtd_actual / mtd_budget * 100) if mtd_budget > 0 else 0
        status_emoji = self.STATUS_EMOJI.get(
            "exceeded" if utilization >= 100 else
            "critical" if utilization >= 90 else
            "warning" if utilization >= 70 else "ok"
        )

        lines = [
            f"📅 *Daily Digest - {entity.upper()}*",
            f"Date: {date}",
            "",
            f"*Today:*",
            f"• Transactions: {transactions_today}",
            f"• Total Spent: {self.CURRENCY_FORMAT.format(total_spent)}",
            "",
            f"*Month-to-Date:*",
            f"• Budget: {self.CURRENCY_FORMAT.format(mtd_budget)}",
            f"• Spent: {self.CURRENCY_FORMAT.format(mtd_actual)}",
            f"• Utilization: {status_emoji} {self.PERCENT_FORMAT.format(utilization)}",
            ""
        ]

        if top_categories:
            lines.append("*Top Categories Today:*")
            for cat, amount in top_categories[:5]:
                lines.append(f"• {cat}: {self.CURRENCY_FORMAT.format(amount)}")

        return "\n".join(lines)
