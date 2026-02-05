"""
Consolidation Engine Module

Generates consolidated financial statements for multiple entities.
Handles inter-company eliminations and group-level reporting.
"""

import json
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import yaml

from .excel_builder import PLExcelBuilder
from .pptx_builder import PLPowerPointBuilder


@dataclass
class ConsolidationEntry:
    """Represents a consolidation adjustment entry."""
    description: str
    debit_account: str
    debit_amount: Decimal
    credit_account: str
    credit_amount: Decimal
    entity_from: str
    entity_to: str
    elimination_type: str  # 'intercompany', 'equity', 'profit'


@dataclass
class ConsolidatedData:
    """Container for consolidated financial data."""
    entities: list[str]
    period: str
    revenue: dict = field(default_factory=dict)
    cost_of_sales: dict = field(default_factory=dict)
    operating_expenses: dict = field(default_factory=dict)
    eliminations: list[ConsolidationEntry] = field(default_factory=list)
    entity_data: dict = field(default_factory=dict)


class ConsolidationEngine:
    """Engine for consolidating financial statements across entities."""

    # Default consolidation groups
    JUNKET_ENTITIES = ["solaire", "cod", "royce"]
    ALL_ENTITIES = ["solaire", "cod", "royce", "manila_junket", "tours", "midori"]

    def __init__(self, config_dir: Path | str = None):
        """Initialize the consolidation engine.

        Args:
            config_dir: Path to configuration directory
        """
        self.config_dir = Path(config_dir) if config_dir else Path(__file__).parent.parent.parent / "config"
        self._load_config()
        self.excel_builder = PLExcelBuilder(config_dir)
        self.pptx_builder = PLPowerPointBuilder(config_dir)

    def _load_config(self) -> None:
        """Load configuration files."""
        with open(self.config_dir / "entity_config.yaml") as f:
            self.entity_config = yaml.safe_load(f)

        with open(self.config_dir / "chart_of_accounts.yaml") as f:
            self.chart_of_accounts = yaml.safe_load(f)

    def consolidate(
        self,
        entities: list[str],
        period: str,
        entity_data: dict[str, dict],
        eliminate_intercompany: bool = True,
    ) -> ConsolidatedData:
        """Consolidate financial data from multiple entities.

        Args:
            entities: List of entity codes to consolidate
            period: Period string (YYYY-MM)
            entity_data: Dictionary of entity financial data
            eliminate_intercompany: Whether to eliminate intercompany transactions

        Returns:
            ConsolidatedData object
        """
        result = ConsolidatedData(
            entities=entities,
            period=period,
            entity_data=entity_data
        )

        # Aggregate revenue
        result.revenue = self._aggregate_section(
            entity_data, entities, "revenue"
        )

        # Aggregate cost of sales
        result.cost_of_sales = self._aggregate_section(
            entity_data, entities, "cost_of_sales"
        )

        # Aggregate operating expenses
        result.operating_expenses = self._aggregate_section(
            entity_data, entities, "operating_expenses"
        )

        # Process intercompany eliminations
        if eliminate_intercompany:
            eliminations = self._calculate_eliminations(entity_data, entities)
            result.eliminations = eliminations
            self._apply_eliminations(result, eliminations)

        return result

    def _aggregate_section(
        self,
        entity_data: dict[str, dict],
        entities: list[str],
        section: str
    ) -> dict:
        """Aggregate a section across entities.

        Args:
            entity_data: Dictionary of entity financial data
            entities: List of entities to aggregate
            section: Section name (revenue, cost_of_sales, operating_expenses)

        Returns:
            Aggregated section data
        """
        aggregated_items = {}

        for entity in entities:
            data = entity_data.get(entity, {})
            section_data = data.get(section, {})
            items = section_data.get("items", [])

            for item in items:
                code = item.get("code", "")
                if code not in aggregated_items:
                    aggregated_items[code] = {
                        "code": code,
                        "name": item.get("name", ""),
                        "actual": Decimal("0"),
                        "budget": Decimal("0"),
                        "prior_period": Decimal("0"),
                        "by_entity": {}
                    }

                actual = Decimal(str(item.get("actual", 0)))
                budget = Decimal(str(item.get("budget", 0)))
                prior = Decimal(str(item.get("prior_period", 0)))

                aggregated_items[code]["actual"] += actual
                aggregated_items[code]["budget"] += budget
                aggregated_items[code]["prior_period"] += prior
                aggregated_items[code]["by_entity"][entity] = {
                    "actual": actual,
                    "budget": budget,
                    "prior_period": prior
                }

        return {
            "items": list(aggregated_items.values()),
            "total_actual": sum(item["actual"] for item in aggregated_items.values()),
            "total_budget": sum(item["budget"] for item in aggregated_items.values()),
        }

    def _calculate_eliminations(
        self,
        entity_data: dict[str, dict],
        entities: list[str]
    ) -> list[ConsolidationEntry]:
        """Calculate intercompany elimination entries.

        Args:
            entity_data: Dictionary of entity financial data
            entities: List of entities being consolidated

        Returns:
            List of elimination entries
        """
        eliminations = []

        # Check for intercompany transactions
        # This is a simplified implementation - in production, you would
        # need to track actual intercompany transactions

        for entity in entities:
            data = entity_data.get(entity, {})
            intercompany = data.get("intercompany", {})

            # Intercompany receivables/payables
            for target_entity, amount in intercompany.get("receivables", {}).items():
                if target_entity in entities:
                    eliminations.append(ConsolidationEntry(
                        description=f"Eliminate intercompany receivable {entity} -> {target_entity}",
                        debit_account="2010",  # Accounts Payable
                        debit_amount=Decimal(str(amount)),
                        credit_account="1030",  # Accounts Receivable
                        credit_amount=Decimal(str(amount)),
                        entity_from=entity,
                        entity_to=target_entity,
                        elimination_type="intercompany"
                    ))

            # Intercompany revenue/expense
            for target_entity, amount in intercompany.get("revenue_from", {}).items():
                if target_entity in entities:
                    eliminations.append(ConsolidationEntry(
                        description=f"Eliminate intercompany revenue {entity} from {target_entity}",
                        debit_account="4050",  # Service Income
                        debit_amount=Decimal(str(amount)),
                        credit_account="6630",  # Consulting Fees (or appropriate expense)
                        credit_amount=Decimal(str(amount)),
                        entity_from=entity,
                        entity_to=target_entity,
                        elimination_type="intercompany"
                    ))

        return eliminations

    def _apply_eliminations(
        self,
        result: ConsolidatedData,
        eliminations: list[ConsolidationEntry]
    ) -> None:
        """Apply elimination entries to consolidated data.

        Args:
            result: ConsolidatedData to modify
            eliminations: List of elimination entries
        """
        for elim in eliminations:
            # Apply to revenue if applicable
            for item in result.revenue.get("items", []):
                if item["code"] == elim.debit_account:
                    item["actual"] -= elim.debit_amount
                    item["elimination"] = item.get("elimination", Decimal("0")) + elim.debit_amount

            # Apply to expenses if applicable
            for section in [result.cost_of_sales, result.operating_expenses]:
                for item in section.get("items", []):
                    if item["code"] == elim.credit_account:
                        item["actual"] -= elim.credit_amount
                        item["elimination"] = item.get("elimination", Decimal("0")) + elim.credit_amount

    def generate_junket_consolidated(
        self,
        period: str,
        entity_data: dict[str, dict],
        output_dir: Path | str,
    ) -> dict[str, Path]:
        """Generate consolidated report for junket entities (Solaire, COD, Royce).

        Args:
            period: Period string (YYYY-MM)
            entity_data: Dictionary of entity financial data
            output_dir: Output directory

        Returns:
            Dictionary of generated file paths
        """
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        # Consolidate data
        consolidated = self.consolidate(
            entities=self.JUNKET_ENTITIES,
            period=period,
            entity_data=entity_data,
            eliminate_intercompany=True
        )

        # Convert to format expected by builders
        report_data = self._to_report_format(consolidated)

        # Generate Excel
        excel_path = output_dir / f"junket_consolidated_pl_{period}.xlsx"
        self.excel_builder.generate_pl_report(
            entity="junket_consolidated",
            period=period,
            data=report_data,
            output_path=excel_path
        )

        # Generate PowerPoint
        pptx_path = output_dir / f"junket_consolidated_pl_{period}.pptx"
        self.pptx_builder.generate_consolidated_report(
            entities=self.JUNKET_ENTITIES,
            period=period,
            data=entity_data,
            output_path=pptx_path
        )

        return {
            "excel": excel_path,
            "pptx": pptx_path,
            "consolidated_data": consolidated
        }

    def generate_group_consolidated(
        self,
        period: str,
        entity_data: dict[str, dict],
        output_dir: Path | str,
    ) -> dict[str, Path]:
        """Generate consolidated report for all entities.

        Args:
            period: Period string (YYYY-MM)
            entity_data: Dictionary of entity financial data
            output_dir: Output directory

        Returns:
            Dictionary of generated file paths
        """
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        # Consolidate data
        consolidated = self.consolidate(
            entities=self.ALL_ENTITIES,
            period=period,
            entity_data=entity_data,
            eliminate_intercompany=True
        )

        # Convert to format expected by builders
        report_data = self._to_report_format(consolidated)

        # Generate Excel
        excel_path = output_dir / f"group_consolidated_pl_{period}.xlsx"
        self.excel_builder.generate_pl_report(
            entity="group_consolidated",
            period=period,
            data=report_data,
            output_path=excel_path
        )

        # Generate PowerPoint
        pptx_path = output_dir / f"group_consolidated_pl_{period}.pptx"
        self.pptx_builder.generate_consolidated_report(
            entities=self.ALL_ENTITIES,
            period=period,
            data=entity_data,
            output_path=pptx_path
        )

        return {
            "excel": excel_path,
            "pptx": pptx_path,
            "consolidated_data": consolidated
        }

    def _to_report_format(self, consolidated: ConsolidatedData) -> dict:
        """Convert ConsolidatedData to report format.

        Args:
            consolidated: ConsolidatedData object

        Returns:
            Dictionary in report format
        """
        def convert_items(items):
            return [{
                "code": item["code"],
                "name": item["name"],
                "actual": float(item["actual"]),
                "budget": float(item["budget"]),
                "prior_period": float(item["prior_period"]),
            } for item in items]

        return {
            "revenue": {
                "items": convert_items(consolidated.revenue.get("items", []))
            },
            "cost_of_sales": {
                "items": convert_items(consolidated.cost_of_sales.get("items", []))
            },
            "operating_expenses": {
                "items": convert_items(consolidated.operating_expenses.get("items", []))
            },
            "eliminations": [
                {
                    "description": e.description,
                    "amount": float(e.debit_amount),
                    "type": e.elimination_type
                }
                for e in consolidated.eliminations
            ]
        }

    def generate_comparison_report(
        self,
        entities: list[str],
        periods: list[str],
        all_period_data: dict[str, dict[str, dict]],
        output_dir: Path | str,
    ) -> Path:
        """Generate a multi-period comparison report.

        Args:
            entities: List of entity codes
            periods: List of period strings to compare
            all_period_data: Nested dict {period: {entity: data}}
            output_dir: Output directory

        Returns:
            Path to generated report
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Period Comparison"

        # Headers
        ws.cell(row=1, column=1, value="Entity")
        ws.cell(row=1, column=2, value="Metric")
        for col, period in enumerate(periods, 3):
            ws.cell(row=1, column=col, value=period)
            ws.cell(row=1, column=col).font = Font(bold=True)

        row = 2
        for entity in entities:
            entity_info = self.entity_config["entities"].get(entity, {})
            entity_name = entity_info.get("name", entity.title())

            # Revenue row
            ws.cell(row=row, column=1, value=entity_name)
            ws.cell(row=row, column=2, value="Revenue")
            for col, period in enumerate(periods, 3):
                period_data = all_period_data.get(period, {}).get(entity, {})
                revenue = sum(
                    Decimal(str(i.get("actual", 0)))
                    for i in period_data.get("revenue", {}).get("items", [])
                )
                ws.cell(row=row, column=col, value=float(revenue))
                ws.cell(row=row, column=col).number_format = '₱#,##0'
            row += 1

            # Net Income row
            ws.cell(row=row, column=1, value="")
            ws.cell(row=row, column=2, value="Net Income")
            for col, period in enumerate(periods, 3):
                period_data = all_period_data.get(period, {}).get(entity, {})
                rev = sum(Decimal(str(i.get("actual", 0))) for i in period_data.get("revenue", {}).get("items", []))
                cos = sum(Decimal(str(i.get("actual", 0))) for i in period_data.get("cost_of_sales", {}).get("items", []))
                opex = sum(Decimal(str(i.get("actual", 0))) for i in period_data.get("operating_expenses", {}).get("items", []))
                net = rev - cos - opex
                ws.cell(row=row, column=col, value=float(net))
                ws.cell(row=row, column=col).number_format = '₱#,##0'
            row += 1

            row += 1  # Blank row between entities

        output_path = output_dir / f"comparison_report_{periods[0]}_{periods[-1]}.xlsx"
        wb.save(output_path)

        return output_path

    def get_consolidation_summary(self, consolidated: ConsolidatedData) -> dict:
        """Get a summary of consolidated results.

        Args:
            consolidated: ConsolidatedData object

        Returns:
            Summary dictionary
        """
        total_revenue = sum(
            item["actual"] for item in consolidated.revenue.get("items", [])
        )
        total_cos = sum(
            item["actual"] for item in consolidated.cost_of_sales.get("items", [])
        )
        total_opex = sum(
            item["actual"] for item in consolidated.operating_expenses.get("items", [])
        )

        gross_profit = total_revenue - total_cos
        net_income = gross_profit - total_opex

        return {
            "period": consolidated.period,
            "entities": consolidated.entities,
            "entity_count": len(consolidated.entities),
            "total_revenue": float(total_revenue),
            "total_cost_of_sales": float(total_cos),
            "gross_profit": float(gross_profit),
            "gross_margin": float(gross_profit / total_revenue * 100) if total_revenue else 0,
            "total_operating_expenses": float(total_opex),
            "net_income": float(net_income),
            "net_margin": float(net_income / total_revenue * 100) if total_revenue else 0,
            "elimination_count": len(consolidated.eliminations),
            "total_eliminations": float(sum(e.debit_amount for e in consolidated.eliminations)),
        }


def main():
    """CLI entry point for testing."""
    import argparse

    parser = argparse.ArgumentParser(description="Generate consolidated reports")
    parser.add_argument("--type", choices=["junket", "group"], required=True, help="Consolidation type")
    parser.add_argument("--month", required=True, help="Period (YYYY-MM)")
    parser.add_argument("--data-dir", required=True, help="Directory with entity JSON files")
    parser.add_argument("--output-dir", default="/tmp/output", help="Output directory")

    args = parser.parse_args()

    # Load entity data
    data_dir = Path(args.data_dir)
    entity_data = {}
    for json_file in data_dir.glob("*.json"):
        entity = json_file.stem.replace(f"_{args.month}", "")
        with open(json_file) as f:
            entity_data[entity] = json.load(f)

    # Generate consolidated report
    engine = ConsolidationEngine()

    if args.type == "junket":
        result = engine.generate_junket_consolidated(
            period=args.month,
            entity_data=entity_data,
            output_dir=args.output_dir
        )
    else:
        result = engine.generate_group_consolidated(
            period=args.month,
            entity_data=entity_data,
            output_dir=args.output_dir
        )

    print(f"Generated Excel: {result['excel']}")
    print(f"Generated PPTX: {result['pptx']}")

    summary = engine.get_consolidation_summary(result['consolidated_data'])
    print(f"\nConsolidation Summary:")
    print(f"  Entities: {summary['entities']}")
    print(f"  Total Revenue: ₱{summary['total_revenue']:,.0f}")
    print(f"  Net Income: ₱{summary['net_income']:,.0f}")
    print(f"  Net Margin: {summary['net_margin']:.1f}%")


if __name__ == "__main__":
    main()
