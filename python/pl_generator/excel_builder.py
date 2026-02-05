"""
P&L Excel Builder Module

Generates formatted Excel P&L reports for all 6 entities using openpyxl.
"""

import json
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


class PLExcelBuilder:
    """Builds formatted P&L Excel reports."""

    # Style definitions
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
    SUBHEADER_FILL = PatternFill(start_color="D6E3F8", end_color="D6E3F8", fill_type="solid")
    SUBHEADER_FONT = Font(bold=True, size=11)
    TOTAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    TOTAL_FONT = Font(bold=True, size=11)
    CURRENCY_FORMAT = '₱#,##0.00'
    PERCENT_FORMAT = '0.0%'

    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def __init__(self, config_dir: Path | str = None):
        """Initialize the Excel builder.

        Args:
            config_dir: Path to configuration directory
        """
        self.config_dir = Path(config_dir) if config_dir else Path(__file__).parent.parent.parent / "config"
        self._load_config()

    def _load_config(self) -> None:
        """Load configuration files."""
        # Load chart of accounts
        with open(self.config_dir / "chart_of_accounts.yaml") as f:
            self.chart_of_accounts = yaml.safe_load(f)

        # Load entity config
        with open(self.config_dir / "entity_config.yaml") as f:
            self.entity_config = yaml.safe_load(f)

    def _setup_styles(self, wb: Workbook) -> None:
        """Set up named styles for the workbook."""
        # Header style
        header_style = NamedStyle(name="header")
        header_style.font = self.HEADER_FONT
        header_style.fill = self.HEADER_FILL
        header_style.alignment = Alignment(horizontal="center", vertical="center")
        if "header" not in wb.named_styles:
            wb.add_named_style(header_style)

        # Currency style
        currency_style = NamedStyle(name="currency")
        currency_style.number_format = self.CURRENCY_FORMAT
        currency_style.alignment = Alignment(horizontal="right")
        if "currency" not in wb.named_styles:
            wb.add_named_style(currency_style)

    def generate_pl_report(
        self,
        entity: str,
        period: str,
        data: dict[str, Any],
        output_path: Path | str,
        template_path: Path | str = None,
        include_budget: bool = True,
        include_variance: bool = True,
        include_prior_period: bool = True,
    ) -> Path:
        """Generate a P&L report for an entity.

        Args:
            entity: Entity code (solaire, cod, etc.)
            period: Period string (YYYY-MM)
            data: Financial data dictionary
            output_path: Output file path
            template_path: Optional template file path
            include_budget: Include budget column
            include_variance: Include variance analysis
            include_prior_period: Include prior period comparison

        Returns:
            Path to generated file
        """
        output_path = Path(output_path)

        # Load template or create new workbook
        if template_path and Path(template_path).exists():
            wb = load_workbook(template_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "P&L Statement"

        self._setup_styles(wb)

        # Get entity info
        entity_info = self.entity_config["entities"].get(entity, {})
        entity_name = entity_info.get("full_name", entity.title())

        # Build the report
        self._build_header(ws, entity_name, period)
        current_row = self._build_revenue_section(ws, data, 5, include_budget, include_prior_period)
        current_row = self._build_cos_section(ws, data, current_row + 1, include_budget, include_prior_period)
        current_row = self._build_gross_profit(ws, data, current_row + 1)
        current_row = self._build_opex_section(ws, data, current_row + 1, include_budget, include_prior_period)
        current_row = self._build_net_income(ws, data, current_row + 1)

        if include_variance:
            current_row = self._build_variance_analysis(ws, data, current_row + 2)

        # Auto-fit columns
        self._auto_fit_columns(ws)

        # Save
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        return output_path

    def _build_header(self, ws: Worksheet, entity_name: str, period: str) -> None:
        """Build the report header."""
        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = f"{entity_name}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")

        # Subtitle
        ws.merge_cells('A2:F2')
        ws['A2'] = "Profit & Loss Statement"
        ws['A2'].font = Font(bold=True, size=14)
        ws['A2'].alignment = Alignment(horizontal="center")

        # Period
        ws.merge_cells('A3:F3')
        period_date = datetime.strptime(period, "%Y-%m")
        ws['A3'] = f"For the Period Ending {period_date.strftime('%B %Y')}"
        ws['A3'].font = Font(italic=True, size=11)
        ws['A3'].alignment = Alignment(horizontal="center")

        # Column headers
        headers = ['Account', 'Description', 'Actual', 'Budget', 'Variance', 'Prior Period']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.style = "header"
            cell.border = self.THIN_BORDER

    def _build_revenue_section(
        self,
        ws: Worksheet,
        data: dict,
        start_row: int,
        include_budget: bool,
        include_prior: bool
    ) -> int:
        """Build the revenue section."""
        row = start_row

        # Section header
        ws.cell(row=row, column=1, value="REVENUE")
        ws.cell(row=row, column=1).font = self.SUBHEADER_FONT
        ws.cell(row=row, column=1).fill = self.SUBHEADER_FILL
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = self.SUBHEADER_FILL
        row += 1

        # Revenue line items
        revenue_data = data.get("revenue", {})
        total_revenue = Decimal("0")
        total_budget = Decimal("0")
        total_prior = Decimal("0")

        for account in revenue_data.get("items", []):
            ws.cell(row=row, column=1, value=account.get("code", ""))
            ws.cell(row=row, column=2, value=account.get("name", ""))

            actual = Decimal(str(account.get("actual", 0)))
            ws.cell(row=row, column=3, value=float(actual))
            ws.cell(row=row, column=3).number_format = self.CURRENCY_FORMAT

            if include_budget:
                budget = Decimal(str(account.get("budget", 0)))
                ws.cell(row=row, column=4, value=float(budget))
                ws.cell(row=row, column=4).number_format = self.CURRENCY_FORMAT
                variance = actual - budget
                ws.cell(row=row, column=5, value=float(variance))
                ws.cell(row=row, column=5).number_format = self.CURRENCY_FORMAT
                total_budget += budget

            if include_prior:
                prior = Decimal(str(account.get("prior_period", 0)))
                ws.cell(row=row, column=6, value=float(prior))
                ws.cell(row=row, column=6).number_format = self.CURRENCY_FORMAT
                total_prior += prior

            total_revenue += actual
            row += 1

        # Total Revenue
        ws.cell(row=row, column=1, value="")
        ws.cell(row=row, column=2, value="Total Revenue")
        ws.cell(row=row, column=2).font = self.TOTAL_FONT
        ws.cell(row=row, column=3, value=float(total_revenue))
        ws.cell(row=row, column=3).number_format = self.CURRENCY_FORMAT
        ws.cell(row=row, column=3).font = self.TOTAL_FONT

        if include_budget:
            ws.cell(row=row, column=4, value=float(total_budget))
            ws.cell(row=row, column=4).number_format = self.CURRENCY_FORMAT
            ws.cell(row=row, column=5, value=float(total_revenue - total_budget))
            ws.cell(row=row, column=5).number_format = self.CURRENCY_FORMAT

        if include_prior:
            ws.cell(row=row, column=6, value=float(total_prior))
            ws.cell(row=row, column=6).number_format = self.CURRENCY_FORMAT

        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = self.TOTAL_FILL
            ws.cell(row=row, column=col).border = self.THIN_BORDER

        return row

    def _build_cos_section(
        self,
        ws: Worksheet,
        data: dict,
        start_row: int,
        include_budget: bool,
        include_prior: bool
    ) -> int:
        """Build the Cost of Sales section."""
        row = start_row

        # Section header
        ws.cell(row=row, column=1, value="COST OF SALES")
        ws.cell(row=row, column=1).font = self.SUBHEADER_FONT
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = self.SUBHEADER_FILL
        row += 1

        # COS line items
        cos_data = data.get("cost_of_sales", {})
        total_cos = Decimal("0")
        total_budget = Decimal("0")
        total_prior = Decimal("0")

        for account in cos_data.get("items", []):
            ws.cell(row=row, column=1, value=account.get("code", ""))
            ws.cell(row=row, column=2, value=account.get("name", ""))

            actual = Decimal(str(account.get("actual", 0)))
            ws.cell(row=row, column=3, value=float(actual))
            ws.cell(row=row, column=3).number_format = self.CURRENCY_FORMAT

            if include_budget:
                budget = Decimal(str(account.get("budget", 0)))
                ws.cell(row=row, column=4, value=float(budget))
                ws.cell(row=row, column=4).number_format = self.CURRENCY_FORMAT
                variance = budget - actual  # For expenses, under budget is positive
                ws.cell(row=row, column=5, value=float(variance))
                ws.cell(row=row, column=5).number_format = self.CURRENCY_FORMAT
                total_budget += budget

            if include_prior:
                prior = Decimal(str(account.get("prior_period", 0)))
                ws.cell(row=row, column=6, value=float(prior))
                ws.cell(row=row, column=6).number_format = self.CURRENCY_FORMAT
                total_prior += prior

            total_cos += actual
            row += 1

        # Total COS
        ws.cell(row=row, column=2, value="Total Cost of Sales")
        ws.cell(row=row, column=2).font = self.TOTAL_FONT
        ws.cell(row=row, column=3, value=float(total_cos))
        ws.cell(row=row, column=3).number_format = self.CURRENCY_FORMAT
        ws.cell(row=row, column=3).font = self.TOTAL_FONT

        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = self.TOTAL_FILL
            ws.cell(row=row, column=col).border = self.THIN_BORDER

        return row

    def _build_gross_profit(self, ws: Worksheet, data: dict, start_row: int) -> int:
        """Build the Gross Profit line."""
        row = start_row

        total_revenue = sum(
            Decimal(str(item.get("actual", 0)))
            for item in data.get("revenue", {}).get("items", [])
        )
        total_cos = sum(
            Decimal(str(item.get("actual", 0)))
            for item in data.get("cost_of_sales", {}).get("items", [])
        )
        gross_profit = total_revenue - total_cos

        ws.cell(row=row, column=2, value="GROSS PROFIT")
        ws.cell(row=row, column=2).font = Font(bold=True, size=12)
        ws.cell(row=row, column=3, value=float(gross_profit))
        ws.cell(row=row, column=3).number_format = self.CURRENCY_FORMAT
        ws.cell(row=row, column=3).font = Font(bold=True, size=12)

        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = PatternFill(
                start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
            )
            ws.cell(row=row, column=col).border = self.THIN_BORDER

        return row

    def _build_opex_section(
        self,
        ws: Worksheet,
        data: dict,
        start_row: int,
        include_budget: bool,
        include_prior: bool
    ) -> int:
        """Build the Operating Expenses section."""
        row = start_row

        # Section header
        ws.cell(row=row, column=1, value="OPERATING EXPENSES")
        ws.cell(row=row, column=1).font = self.SUBHEADER_FONT
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = self.SUBHEADER_FILL
        row += 1

        # OpEx line items
        opex_data = data.get("operating_expenses", {})
        total_opex = Decimal("0")

        for account in opex_data.get("items", []):
            ws.cell(row=row, column=1, value=account.get("code", ""))
            ws.cell(row=row, column=2, value=account.get("name", ""))

            actual = Decimal(str(account.get("actual", 0)))
            ws.cell(row=row, column=3, value=float(actual))
            ws.cell(row=row, column=3).number_format = self.CURRENCY_FORMAT

            if include_budget:
                budget = Decimal(str(account.get("budget", 0)))
                ws.cell(row=row, column=4, value=float(budget))
                ws.cell(row=row, column=4).number_format = self.CURRENCY_FORMAT
                variance = budget - actual
                ws.cell(row=row, column=5, value=float(variance))
                ws.cell(row=row, column=5).number_format = self.CURRENCY_FORMAT

            if include_prior:
                prior = Decimal(str(account.get("prior_period", 0)))
                ws.cell(row=row, column=6, value=float(prior))
                ws.cell(row=row, column=6).number_format = self.CURRENCY_FORMAT

            total_opex += actual
            row += 1

        # Total OpEx
        ws.cell(row=row, column=2, value="Total Operating Expenses")
        ws.cell(row=row, column=2).font = self.TOTAL_FONT
        ws.cell(row=row, column=3, value=float(total_opex))
        ws.cell(row=row, column=3).number_format = self.CURRENCY_FORMAT
        ws.cell(row=row, column=3).font = self.TOTAL_FONT

        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = self.TOTAL_FILL
            ws.cell(row=row, column=col).border = self.THIN_BORDER

        return row

    def _build_net_income(self, ws: Worksheet, data: dict, start_row: int) -> int:
        """Build the Net Income line."""
        row = start_row

        # Calculate net income
        total_revenue = sum(
            Decimal(str(item.get("actual", 0)))
            for item in data.get("revenue", {}).get("items", [])
        )
        total_cos = sum(
            Decimal(str(item.get("actual", 0)))
            for item in data.get("cost_of_sales", {}).get("items", [])
        )
        total_opex = sum(
            Decimal(str(item.get("actual", 0)))
            for item in data.get("operating_expenses", {}).get("items", [])
        )
        net_income = total_revenue - total_cos - total_opex

        ws.cell(row=row, column=2, value="NET INCOME")
        ws.cell(row=row, column=2).font = Font(bold=True, size=14)
        ws.cell(row=row, column=3, value=float(net_income))
        ws.cell(row=row, column=3).number_format = self.CURRENCY_FORMAT
        ws.cell(row=row, column=3).font = Font(bold=True, size=14)

        # Color based on profit/loss
        if net_income >= 0:
            fill_color = "C6EFCE"  # Green for profit
        else:
            fill_color = "FFC7CE"  # Red for loss

        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = PatternFill(
                start_color=fill_color, end_color=fill_color, fill_type="solid"
            )
            ws.cell(row=row, column=col).border = Border(
                left=Side(style='medium'),
                right=Side(style='medium'),
                top=Side(style='medium'),
                bottom=Side(style='medium')
            )

        return row

    def _build_variance_analysis(self, ws: Worksheet, data: dict, start_row: int) -> int:
        """Build a variance analysis section."""
        row = start_row

        ws.merge_cells(f'A{row}:F{row}')
        ws.cell(row=row, column=1, value="VARIANCE ANALYSIS")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1

        # Find significant variances
        all_items = (
            data.get("revenue", {}).get("items", []) +
            data.get("cost_of_sales", {}).get("items", []) +
            data.get("operating_expenses", {}).get("items", [])
        )

        significant_variances = []
        for item in all_items:
            actual = Decimal(str(item.get("actual", 0)))
            budget = Decimal(str(item.get("budget", 0)))
            if budget > 0:
                variance_pct = ((actual - budget) / budget) * 100
                if abs(variance_pct) > 10:  # More than 10% variance
                    significant_variances.append({
                        "code": item.get("code", ""),
                        "name": item.get("name", ""),
                        "actual": actual,
                        "budget": budget,
                        "variance_pct": variance_pct
                    })

        # Sort by absolute variance
        significant_variances.sort(key=lambda x: abs(x["variance_pct"]), reverse=True)

        # Headers
        headers = ['Account', 'Description', 'Actual', 'Budget', 'Variance %', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.border = self.THIN_BORDER
        row += 1

        # Variance items
        for item in significant_variances[:10]:  # Top 10 variances
            ws.cell(row=row, column=1, value=item["code"])
            ws.cell(row=row, column=2, value=item["name"])
            ws.cell(row=row, column=3, value=float(item["actual"]))
            ws.cell(row=row, column=3).number_format = self.CURRENCY_FORMAT
            ws.cell(row=row, column=4, value=float(item["budget"]))
            ws.cell(row=row, column=4).number_format = self.CURRENCY_FORMAT
            ws.cell(row=row, column=5, value=float(item["variance_pct"]) / 100)
            ws.cell(row=row, column=5).number_format = self.PERCENT_FORMAT

            # Status indicator
            if item["variance_pct"] > 20:
                status = "⚠️ Over Budget"
                ws.cell(row=row, column=6).fill = PatternFill(
                    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                )
            elif item["variance_pct"] < -20:
                status = "✓ Under Budget"
                ws.cell(row=row, column=6).fill = PatternFill(
                    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                )
            else:
                status = "~ Within Range"

            ws.cell(row=row, column=6, value=status)
            row += 1

        return row

    def _auto_fit_columns(self, ws: Worksheet) -> None:
        """Auto-fit column widths."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width


def main():
    """CLI entry point for testing."""
    import argparse

    parser = argparse.ArgumentParser(description="Generate P&L Excel report")
    parser.add_argument("--entity", required=True, help="Entity code")
    parser.add_argument("--month", required=True, help="Period (YYYY-MM)")
    parser.add_argument("--data-file", required=True, help="JSON data file")
    parser.add_argument("--output-dir", default="/tmp/output", help="Output directory")

    args = parser.parse_args()

    # Load data
    with open(args.data_file) as f:
        data = json.load(f)

    # Generate report
    builder = PLExcelBuilder()
    output_path = Path(args.output_dir) / f"{args.entity}_pl_{args.month}.xlsx"
    result = builder.generate_pl_report(
        entity=args.entity,
        period=args.month,
        data=data,
        output_path=output_path
    )

    print(f"Generated: {result}")


if __name__ == "__main__":
    main()
